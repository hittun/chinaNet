import openpyxl
import os


#-------------------------------------- 辅助 --------------------------------------#

# 友好显示[][]数据类型
def show(list = []):
    for l in list:
        print(l)

# 筛选数据data中含有items的数据，并返回
def filterbycontains(data = [],items = []):
    result = []
    for row in data:
        ismatch = True
        for item in items:
            if item not in row:
                ismatch = False
        if ismatch:
            result.append(row)
    return result

# 筛选数据data中不含有items的数据，并返回
def filterbynocontains(data = [],items = []):
    result = []
    for row in data:
        ismatch = True
        for item in items:
            if item in row:
                ismatch = False
        if ismatch:
            result.append(row)
    return result



class ExcelUtil:
    __FILENAME = 'result.xlsx'
    wb = None
    ws = None
    iswrited = False
    time = 0
    TIME = 100 # 默认两天数据保存一次

    def __init__(self,filename=None,new=False):
        if filename is not None:
            self.__FILENAME = filename
        if os.path.exists(filename):
            if new:
                self.new()
            else:
                pass
        else:
            self.new()
        self.load()

    def new(self):
        self.wb = openpyxl.Workbook()
        self.wb.save(self.__FILENAME)
        print('newed:' + self.__FILENAME)

    def load(self):
        print('loading:' + self.__FILENAME)
        self.wb = openpyxl.load_workbook(filename=self.__FILENAME, read_only=False, keep_links=True)
        self.ws = self.wb.get_active_sheet()

    # 打印EXCEL数据
    def print(self):
        idx = 0
        for row in self.ws.iter_rows():
            idx = idx + 1
            rowstr = []
            for cell in row:
                rowstr.append(cell.value)
            print(str(idx),rowstr)

    # 返回EXCEL数据
    def getdata(self):
        cols = []
        for row in self.ws.iter_rows():
            rows = []
            for cell in row:
                rows.append(cell.value)
            cols.append(rows)
        return cols

    # 返回行数
    def getrowsnumber(self):
        idx = 0
        for row in self.ws.iter_rows():
            idx = idx + 1
        return idx

    # 返回列数
    def getcolsnumber(self):
        idx = 0
        for col in self.ws.iter_cols():
            idx = idx + 1
        return idx

    # 根据列名字(默认第一行) ,获取特定列数据
    # 注意 ： 能找到就返回，没找到就忽略该item
    def getcolsbyname(self,items=[]):
        colidxs = [] # 用来保存items各个item所在列的索引
        isFirst = True
        result = []
        for row in self.ws.iter_rows():
            if isFirst:  # 第一行
                isFirst = False
                for item in items: # 开始遍历匹配items
                    colidx = 0
                    for cell in row: # 查询 第一行中哪个单元格的值==item , 记录坐标
                        if cell.value == item:
                            colidxs.append(colidx)
                            break
                        colidx = colidx + 1
            else:
                resultitem = []
                for idx in colidxs:
                    resultitem.append(row[idx].value)
                result.append(resultitem)
        return result

    def write(self,items=None):
        self.iswrited = True
        self.ws.append(items)
        # self.time += 1
        # if self.time == self.TIME:
        #     self.time = 0
        #     self.save()
            # self.load()

    def save(self):
        if self.iswrited: # 修改过数据才需要保存
            self.wb.save(filename=self.__FILENAME)
            print('saved:' + self.__FILENAME)

    def close(self):
        self.save() # 最后的数据
        self.wb.close()
        print('closed:'+self.__FILENAME)

    def test(self):
        print('class ExcelUtils Test Sucess')


def test():
    print('ExcelUtils Test Sucess')

if __name__=='__main__':
    pass
