#!/usr/bin/python
# -*- coding: UTF-8 -*-

#@Authod : wuyanhui

'''
多条件匹配筛选
在matchFilters 基础上增加了更安全的保存
'''

import openpyxl
import numpy as np



class ExcelUtil:
    __FILENAME = 'result.xlsx'
    wb = None
    ws = None
    time = 0
    TIME = 100 # 默认两天数据保存一次

    def __init__(self,filename=None):
        if filename is not None:
            self.__FILENAME = filename
        self.new()
        self.load()

    def new(self):
        print('创建文件:'+self.__FILENAME)
        self.wb = openpyxl.Workbook()
        self.wb.save(self.__FILENAME)

    def load(self):
        self.wb = openpyxl.load_workbook(filename=self.__FILENAME, read_only=False, keep_links=True)
        self.ws = self.wb.get_active_sheet()

    def write(self,items=None):
        self.ws.append(items)
        # self.time += 1
        # if self.time == self.TIME:
        #     self.time = 0
        #     self.save()
            # self.load()

    def save(self):
        print('保存数据:' + self.__FILENAME)
        self.wb.save(filename=self.__FILENAME)

    def close(self):
        self.save() # 最后的数据
        self.wb.close()
        print('文件:'+self.__FILENAME+'已关闭')


class MF:
    DATA = None
    FILTER = None
    RESULT = None
    EU = None # 留用

    def __init__(self):
        print('初始化')
        self.DATA = 'data.xlsx'
        self.FILTER = 'filter.xlsx'
        self.RESULT = 'result.xlsx'

    # 获取过滤条件
    def getFilters(self,filename=None):
        if filename is not None:
            self.FILTER = filename
        print('加载筛选文件:'+self.FILTER)
        wb_filter = openpyxl.load_workbook(filename=self.FILTER)
        ws_filter = wb_filter.get_active_sheet()
        filters = []

        for row in ws_filter.iter_rows():
            filter = []
            for cell in row:
                if cell.value is None:
                    continue
                filter.append(str(cell.value))
            filters.append(filter)
        wb_filter.close()
        # print('筛选条件:')
        # print(filters)
        return filters

    # 留用
    # 获取数据集合
    def getDatas(self,filename=None):
        if filename is not None:
            self.DATA = filename
        print('加载数据文件:' + self.DATA)
        wb_data = openpyxl.load_workbook(filename=self.DATA,read_only=True, keep_links=True)
        ws_data = wb_data.get_active_sheet()
        filters = []
        for row in ws_data.iter_rows():
            for cell in row:
                filters.append(str(cell.value))
        wb_data.close()
        pass

    def run(self):
        # 加载过滤条件
        filters = self.getFilters()
        # 先创建一个文件来保存结果吧
        eu = ExcelUtil()
        # 加载数据文件
        print('加载数据文件:' + self.DATA)
        wb_data = openpyxl.load_workbook(filename=self.DATA)
        ws_data = wb_data.get_active_sheet()

        fi = 0
        # 全局筛选每一个数据-条件遍历
        for filter in filters:
            fi = fi + 1
            print(fi)
            # 全局筛选每一个数据-数据遍历
            isMatch = False # 是否有匹配
            for row in ws_data.iter_rows():  # 所有行
                rowValue = []  # 用来保存一行数据
                for cell in row:
                    # 拼凑某行内所有单元格数据
                    rowValue.append(str(cell.value))
                # 如果匹配到数据 - 保存且跳出本次全局匹配
                isMatchRow = True # 匹配该行
                for f in filter:
                    if f not in rowValue:
                        isMatchRow = False
                        break
                if isMatchRow:
                    isMatch = True
                    items = filter+rowValue
                    eu.write(items=items)
                    # print(items)
                    break  # 匹配数据有可能不只一条。牺牲效率。
            if not isMatch: # 如果该filters没匹配到数据
                items = filter + ['无匹配项']
                # items = filter
                eu.write(items=items)
                print(items)

        eu.close()
        wb_data.close()

    def __del__(self):
        print('处理结束')



if __name__=="__main__":

    print("使用说明1")
    mf = MF()
    mf.run()


    exit(0)