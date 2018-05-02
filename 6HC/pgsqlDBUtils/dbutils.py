# -*- coding:utf-8 -*-
import psycopg2
import openpyxl
import math


# 全局变量
S50 = 49
S100 = 98
S200 = 196
S300 = 294
S500 = 490
s50 = 1
s100 = 2
s200 = 4
s300 = 6
s500 = 10

class mine:
    PX = None

    def __init__(self):
        self.PX = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0, 15: 0,
                   16: 0, 17: 0,
                   18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0, 24: 0, 25: 0, 26: 0, 27: 0, 28: 0, 29: 0, 30: 0, 31: 0,
                   32: 0,
                   33: 0, 34: 0, 35: 0, 36: 0, 37: 0, 38: 0, 39: 0, 40: 0, 41: 0, 42: 0, 43: 0, 44: 0, 45: 0, 46: 0,
                   47: 0,
                   48: 0, 49: 0}

#展示 根据索引
def showByIndex(dict={}):
    for i in range(1,50):
        line = str(i)+','+str(dict[i])
        print(line)

#展示 根据迭代器顺序
def showByIterator(dict={}):
    for k, v in dict.items():
        line = str(k)+','+str(v)
        print(line)

# 排序 根据value 排序，value小的键值对放在最前面
def sortdict(dict={}):
    result = {}

    while dict.__len__() != 0:
        isFirst = True
        minv = None
        mink = None
        for k, v in dict.items():
            if isFirst:
                isFirst = False
                mink = k
                minv = v
            if minv > v:
                mink = k
                minv = v
        dict.pop(mink)
        result[mink] = minv
    return result

# 顺序获取dict key
def rankinglist(dict={}):
    list = []
    for k in dict.keys():
        list.append(k)
    return list

# 顺序获取字典 - 排行榜
def rankingdict(list=[]):
    dict = {}
    for i in range(1,50):
        dict[i] = list[i-1]
    return dict

# 查询号码X的rankingdict排名
def searchRanking(dict={},x=None):
    key = None
    for k, v in dict.items():
        if x == v:
            key = k
            break
    return key


# 预测排行榜
def ranking(dict={}):
    return rankingdict(list=rankinglist(dict=sortdict(dict=dict)))


# 回溯预测方案 单次 返回预测排名
def recall(source=None, tema=None):
    sts = Statistics(data=source)

    x50 = sts.getChanceX(x=S50)
    x100 = sts.getChanceX(x=S100)
    x200 = sts.getChanceX(x=S200)
    x300 = sts.getChanceX(x=S300)
    x500 = sts.getChanceX(x=S500)

    cx = mine().PX # 用于保存每位数的权值和
    for i in range(1, 50):
        cx[i] = x50[i] + x100[i] + x200[i] + x300[i] + x500[i]

    _rankiing= ranking(dict=cx) # 给权值和排序最后得到排行榜
    return searchRanking(dict=_rankiing, x=tema) # 返回真实值的预测值(即该数字的预测排名)

# This class provides the functionality we want. You only need to look at
# this if you want to know how this works. It only needs to be defined
# once, no need to muck around with its internals.
class switch(object):
    def __init__(self, value):
        self.value = value
        self.fall = False

    def __iter__(self):
        """Return the match method once, then stop"""
        yield self.match
        raise StopIteration

    def match(self, *args):
        """Indicate whether or not to enter a case suite"""
        if self.fall or not args:
            return True
        elif self.value in args: # changed for v1.5, see below
            self.fall = True
            return True
        else:
            return False

class ExcelUtil:
    __FILENAME = 'result.xlsx'
    wb = None
    ws = None
    time = 0
    TIME = 100 # 默认两天数据保存一次

    def __init__(self,filename=None):
        if filename is not None:
            self.__FILENAME = filename
        # self.new()
        # self.load()

    def new(self):
        print('创建文件:'+self.__FILENAME)
        self.wb = openpyxl.Workbook()
        self.wb.save(self.__FILENAME)

    def load(self):
        self.wb = openpyxl.load_workbook(filename=self.__FILENAME, read_only=False, keep_links=True)
        self.ws = self.wb.get_active_sheet()

    def write(self,items=None):
        self.ws.append(items)
        # self.time += 1
        # if self.time == self.TIME:
        #     self.time = 0
        #     self.save()
            # self.load()

    def save(self):
        print('保存数据:' + self.__FILENAME)
        self.wb.save(filename=self.__FILENAME)

    def close(self):
        self.save() # 最后的数据
        self.wb.close()
        print('文件:'+self.__FILENAME+'已关闭')

    # 获取数据
    def getSheetData(self, filename=None):
        if filename is not None:
            self.__FILENAME = filename
        print('加载文件:' + self.__FILENAME)
        wb = openpyxl.load_workbook(filename=self.__FILENAME)
        ws = wb.get_active_sheet()
        datas = []

        for row in ws.iter_rows():
            data = []
            for cell in row:
                if cell.value is None:
                    continue
                data.append(str(cell.value))
            # data.append(str(row[1].value))
            datas.append(data)
        datas.pop(0)
        wb.close()
        return datas

    # 获取列数据
    def getSheetCols(self, filename=None, col=0):
        if filename is not None:
            self.__FILENAME = filename
        print('加载文件:' + self.__FILENAME)
        wb = openpyxl.load_workbook(filename=self.__FILENAME)
        ws = wb.get_active_sheet()
        cols = []

        for row in ws.iter_rows():
            cols.append(row[col].value)
        cols.pop(0)
        wb.close()
        return cols



class Statistics:
    _DATA = []

    def __init__(self,data=None):
        if data is not None:
            self._DATA = data

    # 统计近X期，每个特码出现的次数
    def getAppearTimes(self,start=0,x=49):
        nearX = self._DATA[start:start + x] # 获得近X期特码
        px = mine().PX
        for k in nearX:
            px[k] = px[k] + 1
        return px

    # 统计近X期，每个特码出现的方差权值
    def getChanceX(self,start=0,x=49):
        cx = mine().PX
        average = round(x / 49)
        at = self.getAppearTimes(start=start, x=x)

        for i in range(1, 50):
            div = at.get(i) - average  # 距离
            if div == 0:
                cx[i] = 0
            elif div > 0:
                cx[i] = pow(div, 2)
            else:
                cx[i] = pow(div, 2) * (-1)

        return cx

    def getChance50X(self, start=0, x=49):
        pass

    def getChance100X(self, start=0, x=98):
        pass

    def getChance200X(self, start=0, x=49):
        pass

    def getChance300X(self, start=0, x=49):
        pass

    def getChance500X(self, start=0, x=49):
        pass




if __name__ == '__main__':

    eu = ExcelUtil(filename='data.xlsx')
    cols = eu.getSheetCols(col=1)  # 所有特码


    tt = 0
    for i in range(0,1000):
        tema = cols.pop(0)
        ri = recall(source=cols,tema=tema)
        if ri>35:
            tt = tt+1
        print(ri)
    # print(tt)


    exit(0)
    # 建立数据库连接
    conn = psycopg2.connect(database='lhc', user='postgres', password='1234', host='127.0.0.1', port='5432')
    cur = conn.cursor()

    sql = """INSERT INTO T_SOURCE(id,tm,date) VALUES(2,1,'2001-07-13')"""
    cur.execute(sql)
    conn.commit()

    # 释放数据库连接
    cur.close()
    conn.close()

